"""XLSX to text converter."""

import logging
from io import BytesIO
from .base import FileConverter, ConversionResult

logger = logging.getLogger(__name__)


class XLSXConverter(FileConverter):
    """Converter for XLSX files to text."""
    
    def can_convert(self, mime_type: str, file_name: str) -> bool:
        """Check if file is an XLSX."""
        return (
            mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or
            file_name.lower().endswith(".xlsx")
        )
    
    async def convert(self, file_content: bytes, file_name: str, mime_type: str) -> ConversionResult:
        """Convert XLSX to text."""
        try:
            import openpyxl
            
            xlsx_file = BytesIO(file_content)
            workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
            
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = [f"[Sheet: {sheet_name}]"]
                
                # Extract data from each row
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Only add non-empty rows
                    if any(cell.strip() for cell in row_data):
                        sheet_text.append(" | ".join(row_data))
                
                if len(sheet_text) > 1:  # More than just the header
                    text_parts.append("\n".join(sheet_text))
            
            text = "\n\n".join(text_parts)
            
            if text.strip():
                logger.info(f"✅ Converted XLSX {file_name} ({len(text)} chars, {len(workbook.sheetnames)} sheets)")
                return ConversionResult(
                    success=True,
                    text=text,
                    metadata={"sheets": len(workbook.sheetnames), "converter": "openpyxl"}
                )
            else:
                return ConversionResult(
                    success=False,
                    text="",
                    error="XLSX file appears to be empty"
                )
                
        except ImportError:
            return ConversionResult(
                success=False,
                text="",
                error="openpyxl library not installed"
            )
        except Exception as e:
            logger.error(f"Error converting XLSX {file_name}: {e}", exc_info=True)
            return ConversionResult(
                success=False,
                text="",
                error=f"Error extracting text from XLSX: {str(e)}"
            )

